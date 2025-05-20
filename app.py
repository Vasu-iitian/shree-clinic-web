from flask import Flask, render_template, request, redirect
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

EXCEL_FILE = 'data.xlsx'

app = Flask(__name__)  # ✅ Define the app here FIRST

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    now = datetime.now()
    date = now.strftime("%d-%m-%Y")
    time = now.strftime("%H:%M:%S")

    data = {
        "Date": date,
        "Time": time,
        "Name": request.form.get("name"),
        "Age": request.form.get("age"),
        "Sex": request.form.get("sex"),
        "Weight": request.form.get("weight"),
        "Mobile": request.form.get("mobile"),
        "Case Type": request.form.get("case"),
        "Diagnosis": request.form.get("diagnosis"),
        "Prescription": request.form.get("prescription"),
        "OPD Type": request.form.get("opd"),
    }

    df = pd.DataFrame([data])

    # Load existing workbook to get last row
    book = load_workbook(EXCEL_FILE)
    sheet = book.active
    start_row = sheet.max_row

    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, startrow=start_row, index=False, header=False)

    return redirect('/')

if __name__ == '__main__':
    app.run(debug=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
